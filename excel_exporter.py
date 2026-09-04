import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


HEADER_FILL = PatternFill(start_color="2C3547", end_color="2C3547", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True, size=9, color="808080")


def export_excel_report(file_path, project_name, inputs, mix_result, batch_info,
                         cost_info, trial_result=None, method_name="ACI 211.1"):
    wb = openpyxl.Workbook()

    # ---------- Sheet 1: Summary/Inputs ----------
    ws = wb.active
    ws.title = "Input Summary"
    _write_header(ws, "Concrete Mix Design Report", project_name, method_name)

    row = 4
    row = _write_section(ws, row, "Design Parameters", [
        ("Target Strength (f'ck)", f"{inputs['fck']} MPa"),
        ("Slump", f"{inputs['slump']} mm"),
        ("Max Aggregate Size", f"{inputs['max_agg_size']} mm"),
        ("Exposure Condition", inputs['exposure'].replace("_", " ").capitalize()),
    ])

    is_bs_or_is = "IS" in method_name or "BS" in method_name
    if is_bs_or_is:
        row = _write_section(ws, row, "", [("Sand Zone", inputs.get('zone', 'II'))], no_title=True)
    else:
        row = _write_section(ws, row, "", [("Fineness Modulus of Sand", inputs['fm_sand'])], no_title=True)

    if "BS" in method_name:
        row = _write_section(ws, row, "", [("Aggregate Type", inputs.get('aggregate_type', 'uncrushed').capitalize())], no_title=True)

    row += 1
    row = _write_section(ws, row, "Aggregate Moisture Correction", [
        ("Fine Aggregate Moisture", f"{inputs.get('fine_moisture', 0)}%"),
        ("Fine Aggregate Absorption", f"{inputs.get('fine_absorption', 0)}%"),
        ("Coarse Aggregate Moisture", f"{inputs.get('coarse_moisture', 0)}%"),
        ("Coarse Aggregate Absorption", f"{inputs.get('coarse_absorption', 0)}%"),
    ])

    row += 1
    row = _write_section(ws, row, "Batch / Site Settings", [
        ("Total Volume Required", f"{inputs.get('volume', 1)} m³"),
        ("Cement Bag Weight", f"{inputs.get('bag_weight', 50)} kg"),
    ])

    if any(float(inputs.get(k, 0) or 0) > 0 for k in ("cement_rate", "fine_rate", "coarse_rate", "water_rate")):
        row += 1
        row = _write_section(ws, row, "Material Rates", [
            ("Cement Rate", f"{inputs.get('cement_rate', 0)} per bag"),
            ("Fine Aggregate Rate", f"{inputs.get('fine_rate', 0)} per kg"),
            ("Coarse Aggregate Rate", f"{inputs.get('coarse_rate', 0)} per kg"),
            ("Water Rate", f"{inputs.get('water_rate', 0)} per liter"),
        ])

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 28

    # ---------- Sheet 2: Mix Design Results ----------
    ws2 = wb.create_sheet("Mix Design Results")
    _write_header(ws2, "Mix Design Results", project_name, method_name)

    row = 4
    ratio_rows = []
    if "target_mean_strength" in mix_result:
        ratio_rows.append(("Target Mean Strength", f'{mix_result["target_mean_strength"]} MPa'))
        ratio_rows.append(("Standard Deviation Used", mix_result["std_deviation"]))
    ratio_rows += [
        ("Slump Category", mix_result["slump_category"]),
        ("W/C Ratio (strength/durability-based)", mix_result["wc_strength"]),
        ("W/C Ratio (exposure limit)", mix_result["wc_limit"]),
        ("Final W/C Ratio Used", mix_result["wc_final"]),
        ("Coarse Aggregate Fraction", mix_result["coarse_fraction"]),
        ("Air Content", f'{mix_result["air_percent"]}%'),
    ]
    if "min_cement_required" in mix_result:
        ratio_rows.append(("Minimum Cement Required", f'{mix_result["min_cement_required"]} kg/m³'))
    row = _write_section(ws2, row, "Design Ratios", ratio_rows)

    row += 1
    row = _write_section(ws2, row, "Batch (Dry / Lab) Quantities — per m³", [
        ("Water", f'{mix_result["water_batch"]} kg'),
        ("Cement", f'{mix_result["cement_batch"]} kg'),
        ("Fine Aggregate", f'{mix_result["fine_batch"]} kg'),
        ("Coarse Aggregate", f'{mix_result["coarse_batch"]} kg'),
    ])

    row += 1
    row = _write_section(ws2, row, "Field (Moisture Adjusted) Quantities — per m³", [
        ("Water", f'{mix_result["water_field"]} kg'),
        ("Cement", f'{mix_result["cement_field"]} kg'),
        ("Fine Aggregate", f'{mix_result["fine_field"]} kg'),
        ("Coarse Aggregate", f'{mix_result["coarse_field"]} kg'),
    ])

    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 22

    # ---------- Sheet 3: Site Batching & Cost ----------
    ws3 = wb.create_sheet("Site Batching & Cost")
    _write_header(ws3, "Site Batching & Cost Estimation", project_name, method_name)

    row = 4
    row = _write_section(ws3, row, "Site Batching", [
        ("Total Volume", f'{batch_info["volume_m3"]} m³'),
        ("Cement Bags per m³", batch_info["bags_per_m3"]),
        ("Water per Bag", f'{batch_info["water_per_bag"]} kg'),
        ("Fine Aggregate per Bag", f'{batch_info["fine_per_bag"]} kg'),
        ("Coarse Aggregate per Bag", f'{batch_info["coarse_per_bag"]} kg'),
        ("Total Cement Bags", batch_info["total_bags"]),
        ("Total Cement", f'{batch_info["total_cement_kg"]} kg'),
        ("Total Water", f'{batch_info["total_water_kg"]} kg'),
        ("Total Fine Aggregate", f'{batch_info["total_fine_kg"]} kg'),
        ("Total Coarse Aggregate", f'{batch_info["total_coarse_kg"]} kg'),
    ])

    if cost_info and cost_info.get("total_cost", 0) > 0:
        row += 1
        row = _write_section(ws3, row, "Cost Estimation", [
            ("Cement Cost", cost_info["cement_cost"]),
            ("Fine Aggregate Cost", cost_info["fine_cost"]),
            ("Coarse Aggregate Cost", cost_info["coarse_cost"]),
            ("Water Cost", cost_info["water_cost"]),
            ("Total Cost", cost_info["total_cost"]),
            ("Cost per m³", cost_info["cost_per_m3"]),
        ])

    ws3.column_dimensions['A'].width = 32
    ws3.column_dimensions['B'].width = 22

    # ---------- Sheet 4: Trial Mix Adjustment (agar computed ho) ----------
    if trial_result:
        ws4 = wb.create_sheet("Trial Mix Adjustment")
        _write_header(ws4, "Trial Mix Adjustment", project_name, method_name)

        row = 4
        _write_section(ws4, row, "Trial Adjustment", [
            ("Target Slump", f'{trial_result["target_slump"]} mm'),
            ("Actual Measured Slump", f'{trial_result["actual_slump"]} mm'),
            ("Slump Difference", f'{trial_result["slump_difference"]} mm'),
            ("Water Correction", f'{trial_result["water_correction"]} kg/m³'),
            ("Adjusted Water", f'{trial_result["adjusted_water"]} kg/m³'),
            ("Adjusted Cement", f'{trial_result["adjusted_cement"]} kg/m³'),
        ])

        ws4.column_dimensions['A'].width = 28
        ws4.column_dimensions['B'].width = 22

    wb.save(file_path)


def _write_header(ws, title, project_name, method_name):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT

    ws["A2"] = f"Project: {project_name or 'Untitled'}"
    ws["A2"].font = SUBTITLE_FONT

    ws["A3"] = f"Method: {method_name}  |  Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A3"].font = SUBTITLE_FONT


def _write_section(ws, start_row, title, rows, no_title=False):
    row = start_row

    if not no_title and title:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12, color="4F8CFF")
        row += 1

    if title or not no_title:
        header_row = row
        ws.cell(row=header_row, column=1, value="Parameter").font = HEADER_FONT
        ws.cell(row=header_row, column=1).fill = HEADER_FILL
        ws.cell(row=header_row, column=2, value="Value").font = HEADER_FONT
        ws.cell(row=header_row, column=2).fill = HEADER_FILL
        row += 1

    for label, value in rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    return row + 1
